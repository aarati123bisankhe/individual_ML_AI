"""Post-processing: stratified time-based splits, sample XLSX workbook
with summary stats and data dictionary, dataset metadata JSON, and README."""

import json
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

OUT  = Path("/home/claude/retail")
DATA = OUT / "data"
SPLITS = OUT / "splits"; SPLITS.mkdir(exist_ok=True)
SAMPLES = OUT / "samples"; SAMPLES.mkdir(exist_ok=True)

txn_df = pd.read_csv(DATA / "transactions.csv", parse_dates=["entry_timestamp", "exit_timestamp"])
prods  = pd.read_csv(DATA / "products.csv")
custs  = pd.read_csv(DATA / "customers.csv")
stores = pd.read_csv(DATA / "stores.csv")
zones  = pd.read_csv(DATA / "zones.csv")
aisles = pd.read_csv(DATA / "aisles.csv")

# ---------------------------------------------------------------------------
# 1. Time-based splits (train: oldest 70%, val: next 15%, test: newest 15%)
#    This is the right way to split retail data — chronological,
#    not random — to mirror how a real production model is trained.
# ---------------------------------------------------------------------------
txn_sorted = txn_df.sort_values("entry_timestamp").reset_index(drop=True)
n = len(txn_sorted)
n_train = int(n * 0.70)
n_val   = int(n * 0.15)

train_ids = txn_sorted.iloc[:n_train]["transaction_id"]
val_ids   = txn_sorted.iloc[n_train:n_train + n_val]["transaction_id"]
test_ids  = txn_sorted.iloc[n_train + n_val:]["transaction_id"]

train_ids.to_csv(SPLITS / "train_transaction_ids.csv", index=False, header=True)
val_ids.to_csv(SPLITS / "val_transaction_ids.csv", index=False, header=True)
test_ids.to_csv(SPLITS / "test_transaction_ids.csv", index=False, header=True)

split_boundaries = {
    "train_start": str(txn_sorted.iloc[0]["entry_timestamp"]),
    "train_end":   str(txn_sorted.iloc[n_train - 1]["entry_timestamp"]),
    "val_start":   str(txn_sorted.iloc[n_train]["entry_timestamp"]),
    "val_end":     str(txn_sorted.iloc[n_train + n_val - 1]["entry_timestamp"]),
    "test_start":  str(txn_sorted.iloc[n_train + n_val]["entry_timestamp"]),
    "test_end":    str(txn_sorted.iloc[-1]["entry_timestamp"]),
    "n_train": int(n_train),
    "n_val":   int(n_val),
    "n_test":  int(n - n_train - n_val),
}
print("Splits:", split_boundaries)

# ---------------------------------------------------------------------------
# 2. Sample XLSX workbook
#    Bulk data stays in CSVs; the workbook contains a 1,000-row preview,
#    summary statistics, distributions, and a data dictionary.
# ---------------------------------------------------------------------------
wb = Workbook()

HEADER_FILL = PatternFill("solid", start_color="1F4E78")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
BODY_FONT   = Font(name="Arial", size=10)
THIN        = Side(border_style="thin", color="BFBFBF")
BORDER      = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP        = Alignment(wrap_text=True, vertical="top")


def style_header(ws, row=1):
    for cell in ws[row]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER


def autosize(ws, max_w=55):
    for col_cells in ws.columns:
        letter = col_cells[0].column_letter
        m = 0
        for c in col_cells:
            if c.value is None: continue
            l = min(len(str(c.value)), max_w)
            if l > m: m = l
        ws.column_dimensions[letter].width = max(10, min(m + 2, max_w))


# README sheet ---------------------------------------------------------------
ws = wb.active; ws.title = "README"
ws.append(["Field", "Description"])
style_header(ws)

readme_rows = [
    ["Dataset",       "UrbanMart Smart Retail (Synthetic)"],
    ["Version",       "1.0.0"],
    ["Created",       datetime.now().strftime("%Y-%m-%d")],
    ["Project",       "Smart Grocery Navigation + Scan-to-Pay"],
    ["Modeled on",    "Generic Kathmandu-valley urban supermarket chain — no real company is named or claimed"],
    ["Sample size",   f"{len(txn_df):,} transactions / 652,655 line items / 158,241 app events / 52,879 payment events"],
    ["Customers",     "25,000 unique customer profiles with loyalty tier, demographics and home area"],
    ["Stores",        "8 stores across Kathmandu valley (Maharajgunj, Pulchowk, Chuchhepati, Bhaktapur, Naxal, Sanepa, Tinkune, Budhanilkantha)"],
    ["Zones / aisles","17 zones, 200 aisles total across all stores; every product carries (zone, aisle, shelf) coordinates"],
    ["Date range",    f"{txn_df['entry_timestamp'].min()} → {txn_df['entry_timestamp'].max()}"],
    ["Currency",      "Nepalese Rupee (NPR), VAT 13%"],
    ["Payment rails", "Cash, eSewa, Khalti, IME Pay, FonePay, ConnectIPS, Card"],
    ["Splits",        "Time-based 70/15/15 — see splits/ folder"],
    ["License",       "Synthetic — free to use in thesis, code repos, derived works"],
    ["Generator",     "generate_dataset.py (seed=7, fully reproducible, parameterized)"],
    ["Scale-up",      "Re-run with N_TRANSACTIONS=100_000 for the originally requested scale (~3-4 min on a normal laptop)"],
]
for r in readme_rows:
    ws.append(r)
ws.column_dimensions["A"].width = 22
ws.column_dimensions["B"].width = 110
for row in ws.iter_rows(min_row=2):
    for c in row:
        c.font = BODY_FONT
        c.alignment = WRAP


# Data dictionary sheet ------------------------------------------------------
ws = wb.create_sheet("Data Dictionary")
ws.append(["Table (CSV)", "Column", "Type", "Description"])
style_header(ws)

DD = [
    # stores
    ("stores.csv", "store_id",       "string",  "Primary key. e.g. S01."),
    ("stores.csv", "store_name",     "string",  "Display name."),
    ("stores.csv", "city",           "string",  "Kathmandu / Lalitpur / Bhaktapur."),
    ("stores.csv", "area",           "string",  "Neighborhood within the city."),
    ("stores.csv", "latitude",       "float",   "Approximate decimal latitude."),
    ("stores.csv", "longitude",      "float",   "Approximate decimal longitude."),
    ("stores.csv", "size_sqft",      "int",     "Floor area in square feet."),
    ("stores.csv", "num_aisles",     "int",     "Total aisle count in the store."),
    ("stores.csv", "opening_year",   "int",     "Year the store opened."),
    # zones
    ("zones.csv",  "zone_id",        "string",  "Primary key. e.g. Z02 (Fresh Produce)."),
    ("zones.csv",  "zone_name",      "string",  "Human-readable zone label."),
    ("zones.csv",  "ambient_temp_c", "int",     "Ambient temperature setpoint in °C."),
    ("zones.csv",  "is_perimeter",   "bool",    "Whether the zone is on the store perimeter (entry, fresh, dairy, frozen, meat, checkout)."),
    # aisles
    ("aisles.csv", "aisle_id",       "string",  "Primary key. e.g. A0042."),
    ("aisles.csv", "store_id",       "FK",      "Which store this aisle belongs to."),
    ("aisles.csv", "zone_id",        "FK",      "Which zone this aisle sits in."),
    ("aisles.csv", "aisle_number_in_zone","int","Sequential index within the zone."),
    ("aisles.csv", "x_coord",        "float",   "Floor X coordinate (0-100 relative)."),
    ("aisles.csv", "y_coord",        "float",   "Floor Y coordinate (0-100 relative)."),
    # products
    ("products.csv","product_id",    "string",  "Primary key. e.g. P001234."),
    ("products.csv","sku",           "string",  "Stock-keeping unit string."),
    ("products.csv","product_name",  "string",  "Display name (e.g. 'Wai Wai Chicken (Pack of 30)')."),
    ("products.csv","brand",         "string",  "Brand."),
    ("products.csv","subcategory",   "string",  "e.g. Vegetables, Snacks, Personal."),
    ("products.csv","zone_id",       "FK",      "Zone this SKU is shelved in."),
    ("products.csv","aisle_id",      "FK",      "Aisle this SKU is shelved in."),
    ("products.csv","shelf_number",  "int",     "Shelf level within the aisle (1-6)."),
    ("products.csv","shelf_position","int",     "Facing position within the shelf (1-12)."),
    ("products.csv","store_id",      "FK",      "Store carrying this SKU listing (each SKU has one row per store)."),
    ("products.csv","unit",          "string",  "kg / pack / pc / bottle / etc."),
    ("products.csv","price_npr",     "int",     "Listed unit price in Nepalese Rupees."),
    # customers
    ("customers.csv","customer_id",  "string",  "Primary key. e.g. C0001234."),
    ("customers.csv","loyalty_tier", "string",  "Standard / Silver / Gold / Platinum."),
    ("customers.csv","gender",       "string",  "F / M / Other."),
    ("customers.csv","age_bucket",   "string",  "18-24, 25-34, 35-44, 45-54, 55-64, 65+."),
    ("customers.csv","home_area",    "string",  "Neighborhood — drives nearest-store assignment."),
    ("customers.csv","registered_on","date",    "Loyalty signup date."),
    ("customers.csv","uses_app",     "bool",    "Whether the customer uses the smart-shopping app (~62% true)."),
    # transactions
    ("transactions.csv","transaction_id","string","Primary key. e.g. T00000001."),
    ("transactions.csv","customer_id","FK",     "Buyer."),
    ("transactions.csv","store_id",  "FK",      "Store of purchase."),
    ("transactions.csv","entry_timestamp","datetime","When the customer entered the store / app session began."),
    ("transactions.csv","exit_timestamp","datetime","When the customer exited / paid."),
    ("transactions.csv","dwell_minutes","float", "Time spent in store."),
    ("transactions.csv","num_distinct_items","int","Distinct SKUs in basket."),
    ("transactions.csv","num_total_units","int","Sum of all unit quantities."),
    ("transactions.csv","aisles_visited","int", "Distinct aisles touched."),
    ("transactions.csv","n_app_searches","int", "In-app product searches during this trip."),
    ("transactions.csv","n_app_directions","int","In-app 'navigate to aisle' requests during this trip."),
    ("transactions.csv","subtotal_npr","int",   "Sum of line totals before discount/VAT."),
    ("transactions.csv","discount_npr","int",   "Loyalty + festival discount."),
    ("transactions.csv","total_after_discount_npr","int","Subtotal − discount."),
    ("transactions.csv","vat_13pct_npr","int",  "13% VAT (Nepali standard)."),
    ("transactions.csv","grand_total_npr","int","Final amount the customer paid."),
    ("transactions.csv","payment_method","string","Cash / eSewa / Khalti / IME Pay / FonePay / ConnectIPS / Card."),
    ("transactions.csv","scan_to_pay_used","bool","Whether the trip used the in-app scan-to-pay flow."),
    ("transactions.csv","payment_attempts","int","1-3 — multi-attempt payments simulate retries on digital rails."),
    ("transactions.csv","payment_status","string","success / failed."),
    ("transactions.csv","customer_loyalty_tier","string","Tier at time of transaction."),
    ("transactions.csv","is_festival_period","bool","Dashain / Tihar / Nepali New Year window."),
    # transaction_items (THE big table)
    ("transaction_items.csv","line_item_id","string","Primary key. e.g. L000000001."),
    ("transaction_items.csv","transaction_id","FK","Parent transaction."),
    ("transaction_items.csv","product_id","FK","Product SKU at this store."),
    ("transaction_items.csv","sku","string","Stock-keeping unit."),
    ("transaction_items.csv","product_name","string","Denormalized display name."),
    ("transaction_items.csv","zone_id","FK","Zone where the item was located (denormalized for routing analyses)."),
    ("transaction_items.csv","aisle_id","FK","Aisle where the item was located."),
    ("transaction_items.csv","quantity","int","How many units were scanned."),
    ("transaction_items.csv","unit_price_npr","int","Per-unit price."),
    ("transaction_items.csv","line_total_npr","int","quantity × unit_price."),
    ("transaction_items.csv","scan_timestamp","datetime","When this specific item was scanned."),
    ("transaction_items.csv","scan_method","string","barcode / qr_code / manual_entry (app users) or checkout_pos (cashier-scanned)."),
    ("transaction_items.csv","scan_duration_ms","int","Milliseconds from camera-open to confirm-add."),
    ("transaction_items.csv","rescans","int","0–2 — how many times the item had to be re-scanned (model fraud/error signal)."),
    ("transaction_items.csv","scan_success","bool","Whether the scan was eventually successful."),
    # app_events
    ("app_events.csv","event_id","string","Primary key."),
    ("app_events.csv","transaction_id","FK","Trip this event belongs to."),
    ("app_events.csv","customer_id","FK","Acting customer."),
    ("app_events.csv","store_id","FK","Where the event occurred."),
    ("app_events.csv","event_timestamp","datetime","When the event happened."),
    ("app_events.csv","event_type","string","search_query | directions_request."),
    ("app_events.csv","search_query","string","Free-text query (uses real product names so search→product is learnable)."),
    ("app_events.csv","target_aisle_id","FK","For directions_request events — the aisle the user asked to navigate to."),
    # payment_events
    ("payment_events.csv","payment_event_id","string","Primary key."),
    ("payment_events.csv","transaction_id","FK","Parent transaction."),
    ("payment_events.csv","attempt_number","int","1, 2, or 3."),
    ("payment_events.csv","payment_method","string","Same vocabulary as transactions.payment_method."),
    ("payment_events.csv","attempt_timestamp","datetime","When this attempt was made."),
    ("payment_events.csv","latency_ms","int","Round-trip processing time."),
    ("payment_events.csv","status","string","success / failed / retry."),
    ("payment_events.csv","amount_npr","int","Amount of this specific attempt (0 for failed/retry)."),
]
for r in DD:
    ws.append(r)
style_header(ws)
for row in ws.iter_rows(min_row=2):
    for c in row:
        c.font = BODY_FONT
        c.alignment = WRAP
ws.column_dimensions["A"].width = 22
ws.column_dimensions["B"].width = 22
ws.column_dimensions["C"].width = 12
ws.column_dimensions["D"].width = 90


# Summary stats sheet --------------------------------------------------------
ws = wb.create_sheet("Summary Stats")
ws.append(["Metric", "Value"])
style_header(ws)

avg_basket = round(txn_df["num_distinct_items"].mean(), 2)
avg_dwell  = round(txn_df["dwell_minutes"].mean(), 2)
avg_basket_value = round(txn_df["grand_total_npr"].mean(), 2)
total_revenue = int(txn_df["grand_total_npr"].sum())

stats = [
    ["Total transactions", len(txn_df)],
    ["Total line items",   652_655],
    ["Total customers",    25_000],
    ["Active customers (≥1 txn)", txn_df["customer_id"].nunique()],
    ["Avg basket size (items)", avg_basket],
    ["Avg dwell time (min)", avg_dwell],
    ["Avg basket value (NPR)", avg_basket_value],
    ["Total simulated revenue (NPR)", total_revenue],
    ["Scan-to-pay adoption %", round(100 * txn_df["scan_to_pay_used"].mean(), 2)],
    ["Festival-period transactions %", round(100 * txn_df["is_festival_period"].mean(), 2)],
    ["Payment success %", round(100 * (txn_df["payment_status"] == "success").mean(), 2)],
]
for r in stats:
    ws.append(r)

ws.append([]); ws.append(["Payment method distribution"])
for k, v in txn_df["payment_method"].value_counts(normalize=True).round(4).items():
    ws.append([k, v])

ws.append([]); ws.append(["Loyalty tier distribution (transactions)"])
for k, v in txn_df["customer_loyalty_tier"].value_counts(normalize=True).round(4).items():
    ws.append([k, v])

ws.append([]); ws.append(["Store traffic share"])
for k, v in (txn_df["store_id"].value_counts(normalize=True).round(4)).items():
    ws.append([k, v])

ws.append([]); ws.append(["Time-based split boundaries"])
for k, v in split_boundaries.items():
    ws.append([k, str(v)])

autosize(ws)
for row in ws.iter_rows(min_row=2):
    for c in row:
        if c.font != HEADER_FONT:
            c.font = BODY_FONT


# Sample data sheets (1000 rows each) ----------------------------------------
def add_sample_sheet(name, df, n=1000):
    ws = wb.create_sheet(name)
    sample = df.head(n)
    for r in dataframe_to_rows(sample, index=False, header=True):
        ws.append(r)
    style_header(ws)
    ws.freeze_panes = "A2"
    autosize(ws, max_w=40)


add_sample_sheet("Stores",        stores)
add_sample_sheet("Zones",         zones)
add_sample_sheet("Aisles (sample)", aisles, n=200)
add_sample_sheet("Products (sample)", prods, n=500)
add_sample_sheet("Customers (sample)", custs, n=500)
add_sample_sheet("Transactions (sample)", txn_df, n=1000)
add_sample_sheet("Line Items (sample)",
                 pd.read_csv(DATA / "transaction_items.csv", nrows=1000), n=1000)
add_sample_sheet("App Events (sample)",
                 pd.read_csv(DATA / "app_events.csv", nrows=500), n=500)
add_sample_sheet("Payment Events (sample)",
                 pd.read_csv(DATA / "payment_events.csv", nrows=500), n=500)

xlsx_path = SAMPLES / "smart_retail_dataset_overview.xlsx"
wb.save(xlsx_path)
print("Saved", xlsx_path)
print("Sheets:", wb.sheetnames)


# ---------------------------------------------------------------------------
# 3. dataset_metadata.json
# ---------------------------------------------------------------------------
metadata = {
    "dataset_name": "UrbanMart Smart Retail (Synthetic)",
    "version": "1.0.0",
    "created": datetime.now().strftime("%Y-%m-%d"),
    "project": "Smart Grocery Navigation + Scan-to-Pay thesis",
    "modeled_on": "Generic Kathmandu-valley urban supermarket chain (no real company named or claimed)",
    "sample_sizes": {
        "stores": 8,
        "zones": 17,
        "aisles": 200,
        "products": 1168,
        "customers": 25000,
        "transactions": int(len(txn_df)),
        "line_items": 652655,
        "app_events": 158241,
        "payment_events": 52879,
    },
    "currency": "NPR",
    "vat_pct": 13,
    "payment_methods": ["Cash","eSewa","Khalti","IME Pay","FonePay","ConnectIPS","Card"],
    "scan_methods": ["barcode","qr_code","manual_entry","checkout_pos"],
    "loyalty_tiers": ["Standard","Silver","Gold","Platinum"],
    "festivals_modeled": ["Dashain (Oct 5-16)","Tihar (Nov 1-7)","Nepali New Year (Apr 10-16)"],
    "splits": split_boundaries,
    "files": {
        "data/stores.csv": "8 store metadata rows.",
        "data/zones.csv": "17 zone definitions with ambient temperature.",
        "data/aisles.csv": "200 aisle-level records with x/y floor coordinates.",
        "data/products.csv": "Per-store product listings (~1,168 rows, each carrying shelf coordinates).",
        "data/customers.csv": "25,000 anonymized customer profiles.",
        "data/transactions.csv": "Transaction headers (50,000 rows).",
        "data/transaction_items.csv": "Line items with scan event data (~653K rows).",
        "data/app_events.csv": "In-app search and navigation events (~158K rows).",
        "data/payment_events.csv": "Per-attempt payment events (~53K rows).",
        "splits/train_transaction_ids.csv": "70% earliest transactions (chronological train set).",
        "splits/val_transaction_ids.csv":   "Next 15% (chronological validation set).",
        "splits/test_transaction_ids.csv":  "Newest 15% (chronological test set).",
        "samples/smart_retail_dataset_overview.xlsx": "Multi-sheet workbook: README, data dictionary, summary stats, 1000-row previews of every table.",
        "generate_dataset.py": "Reproducible generator (seed=7, parameterized N_TRANSACTIONS).",
    },
    "research_tasks_supported": [
        "Smart navigation: search-query → target_aisle prediction (using app_events + products).",
        "Shelf placement / store layout optimization via market-basket + co-aisle co-occurrence.",
        "In-store path / dwell prediction using zone & aisle visits per trip.",
        "Scan-to-pay fraud / error detection (rescans, scan_success, payment retries).",
        "Checkout latency / payment-rail performance benchmarking (latency_ms, payment_attempts).",
        "Customer segmentation (loyalty × demographics × app usage).",
        "Demand forecasting at SKU × store × day granularity.",
        "Festival surge analysis using is_festival_period flag.",
        "Recommendation systems — next-best-product given current basket.",
    ],
    "limitations": [
        "Synthetic. Statistical structure approximates an urban Nepali supermarket but does not represent any real chain.",
        "Co-purchase patterns are theme-template-driven, not learned from real receipts — basket associations are reasonable but not perfectly realistic.",
        "Generator was run at 50K transactions (~653K line items) to fit the build environment; reset N_TRANSACTIONS=100_000 in generate_dataset.py to scale up to the originally requested size.",
        "Spatial coordinates in aisles.csv are randomly placed within a 0-100 unit floor; for a real store-routing study, replace these with a true store layout.",
        "No PII; no scraping; safe for thesis publication.",
    ],
}
(OUT / "dataset_metadata.json").write_text(json.dumps(metadata, indent=2, ensure_ascii=False))
print("Wrote dataset_metadata.json")
